# -*- coding: utf-8 -*-
from DrissionPage import ChromiumPage, ChromiumOptions
import pandas as pd
import time
import os
import re
from bs4 import BeautifulSoup
from datetime import datetime
from openpyxl import load_workbook, Workbook
from urllib.parse import urlencode, quote
import sys
import shutil
import math

# ================= 🔧 配置区域 =================
API_URL_BASE = "https://shandong.chinatax.gov.cn/module/web/jpage/dataproxy.jsp"
HOME_URL = "https://shandong.chinatax.gov.cn/col/col1053/index.html?number=A0301"
BASE_URL = "https://shandong.chinatax.gov.cn"

COLUMN_ID = 1053
UNIT_ID = 48166

# 文件名
FILE_NAME = "山东税务_全量数据.xlsx"
VERSION = "v21.0 (直接导航 + 双重分页参数)"


# ================= 📂 自动化文件管理 =================

def get_desktop_path():
    """直接获取桌面路径"""
    return os.path.join(os.path.expanduser("~"), "Desktop", FILE_NAME)


def init_or_check_excel(filepath):
    print(f"\n📂 目标文件: {filepath}")

    if os.path.exists(filepath):
        print("✅ [检测结果] 文件已存在！")
        print("   -> 模式：【断点续传】(自动跳过旧数据)")
        # 自动备份
        try:
            bak_path = filepath + ".bak"
            shutil.copy(filepath, bak_path)
        except:
            pass
    else:
        print("🆕 [检测结果] 文件不存在。")
        print("   -> 模式：【全新抓取】")
        wb = Workbook()
        ws = wb.active
        ws.append(["标题", "发文机构", "发文字号", "发文日期", "有效性", "是否涉税法律", "正文内容", "链接"])
        wb.save(filepath)


def get_history_links(filepath):
    """读取历史链接"""
    if not os.path.exists(filepath): return set()
    try:
        df = pd.read_excel(filepath, engine="openpyxl", usecols=["链接"])
        return set(df["链接"].dropna().astype(str).tolist())
    except:
        return set()


def save_row_immediately(row_data, filepath):
    """实时写入"""
    try:
        wb = load_workbook(filepath)
        ws = wb.active
        ws.append(list(row_data.values()))
        wb.save(filepath)
        print(".", end="", flush=True)
    except PermissionError:
        print(f"\n🚨 [严重] 文件被占用！请关闭桌面的 Excel 文件！")
    except Exception as e:
        print(f"\n❌ 写入失败: {e}")


# ================= 🧠 提取逻辑 =================
def safe_re_extract(pattern, text):
    try:
        m = re.search(pattern, text, re.DOTALL)
        if m: return m.group(1).strip()
    except:
        pass
    return ""


def extract_detail(page, url):
    try:
        # 访问详情页
        page.get(url, timeout=10)

        # 遇到防火墙等待
        if "安全检查" in page.title:
            print(f"\n⚠️ 遭遇防火墙: {url}")
            time.sleep(5)

        html = page.html
        soup = BeautifulSoup(html, 'html.parser')

        info = {
            "标题": "", "发文机构": "", "发文字号": "",
            "发文日期": "", "有效性": "未注明",
            "是否涉税法律": "未注明",
            "正文内容": "", "链接": url
        }

        # 1. 源码提取 (最稳)
        if not info['发文机构']:
            for tag in ['发文机关', '发布机构', '发文单位']:
                val = safe_re_extract(f'(.*?)', html)
                if val:
                    info['发文机构'] = val
                    break
        if not info['发文字号']:
            info['发文字号'] = safe_re_extract(r'(.*?)', html)
        if not info['发文日期']:
            for tag in ['发文日期', '发布日期', '成文日期']:
                val = safe_re_extract(f'(.*?)', html)
                if val:
                    info['发文日期'] = val
                    break
        if not info['标题']:
            info['标题'] = safe_re_extract(r'(.*?)', html)

        # 2. 表格补救
        try:
            meta_table = soup.find('table', id='xxgkbg')
            if meta_table:
                tds = meta_table.find_all('td')
                for i, td in enumerate(tds):
                    txt = td.get_text(strip=True)
                    if not info['发文机构'] and ('发文机关' in txt or '发布机构' in txt) and i + 1 < len(tds):
                        info['发文机构'] = tds[i + 1].get_text(strip=True)
                    if not info['发文字号'] and '发文字号' in txt and i + 1 < len(tds):
                        info['发文字号'] = tds[i + 1].get_text(strip=True)
                    if not info['发文日期'] and ('日期' in txt) and i + 1 < len(tds):
                        info['发文日期'] = tds[i + 1].get_text(strip=True)
                    if not info['有效性'] and '有效性' in txt and i + 1 < len(tds):
                        info['有效性'] = tds[i + 1].get_text(strip=True)
                    if '是否涉税法律' in txt and i + 1 < len(tds):
                        info['是否涉税法律'] = tds[i + 1].get_text(strip=True)
        except:
            pass

        # 3. 文本补救
        if not info['发文日期']:
            main_div = soup.find('div', class_='main_content')
            if main_div:
                val = safe_re_extract(r"日期[：:]\s*(\d{4}-\d{2}-\d{2})", main_div.get_text())
                if val: info['发文日期'] = val
        if info['有效性'] == "未注明":
            main_div = soup.find('div', class_='main_content')
            if main_div:
                val = safe_re_extract(r"有效性[：:]\s*(.*?)(?:\s|$)", main_div.get_text())
                if val: info['有效性'] = val
        if not info['标题']:
            t = soup.find('meta', attrs={'name': 'ArticleTitle'})
            if t: info['标题'] = t.get('content', '')

        # 正文
        content_div = soup.find(id='zoom') or soup.find(class_='TRS_Editor')
        if content_div:
            info['正文内容'] = content_div.get_text(strip=True)[:30000]
        else:
            div3 = soup.find('div', class_='main_content3')
            if div3: info['正文内容'] = div3.get_text(strip=True)[:30000]

        print(f"  [ok] {info['标题'][:10]}... | 涉税:{info['是否涉税法律']}")
        return info

    except Exception as e:
        print(f"\n    ❌ 详情错误: {e}")
        return None


# ================= 🚀 主程序 =================
def main():
    print(f"🚀 启动采集器 - {VERSION}")

    # 1. 自动获取桌面路径
    save_path = get_desktop_path()

    # 2. 初始化检查
    init_or_check_excel(save_path)

    # 3. 读取断点
    processed_urls = get_history_links(save_path)
    print(f"📚 历史记录: {len(processed_urls)} 条 (将自动跳过)")

    # 4. 浏览器
    co = ChromiumOptions()
    co.set_user_agent(
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36')
    co.set_argument('--blink-settings=imagesEnabled=false')
    co.ignore_certificate_errors()
    page = ChromiumPage(addr_or_opts=co)

    print(f"🌐 初始化: {HOME_URL}")
    page.get(HOME_URL)
    time.sleep(2)

    BATCH_SIZE = 45

    # 抓取循环 (从第1条到第3000条)
    for start_rec in range(1, 3000, BATCH_SIZE):
        end_rec = start_rec + BATCH_SIZE - 1

        # 【核心修复】计算页码：TRS系统有时依赖page参数
        # start=1 -> page=1, start=46 -> page=2
        page_num = math.ceil(start_rec / BATCH_SIZE)

        print(f"\n🔄 请求区间: {start_rec} - {end_rec} (第 {page_num} 页)")

        # 【核心修复】构造完整的 URL，直接让浏览器跳转过去！
        # 包含了所有可能的参数，确保分页生效
        params = {
            "col": "1",
            "appid": "1",
            "webid": "1",
            "path": "/",
            "columnid": str(COLUMN_ID),
            "unitid": str(UNIT_ID),
            "webname": "国家税务总局山东省税务局",
            "permissiontype": "0",
            "page": str(page_num),  # 显式指定页码
            "startrecord": str(start_rec),  # 显式指定起始行
            "endrecord": str(end_rec)  # 显式指定结束行
        }

        full_api_url = f"{API_URL_BASE}?{urlencode(params)}"

        # 让浏览器直接访问 XML 接口
        page.get(full_api_url)
        xml_text = page.html  # 获取页面内容

        if not xml_text or "wzws" in xml_text:
            print("⚠️ 防火墙拦截，暂停5秒...")
            time.sleep(5)
            continue

        pattern = r'<record><!\[CDATA\[(.*?)\]\]></record>'
        matches = re.findall(pattern, xml_text, re.DOTALL)

        if not matches:
            print(f"🏁 本页无数据 (抓取结束)。")
            break

        print(f"   📄 发现 {len(matches)} 条")

        # 检查是否依然返回 300 条
        if len(matches) > BATCH_SIZE + 20:
            print("⚠️ 严重警告：服务器依然返回全部数据（分页彻底失效）。")
            print("   -> 正在启动【强制跳过】模式，直到找到新数据为止...")

        new_count = 0
        for html_snippet in matches:
            soup = BeautifulSoup(html_snippet, 'html.parser')
            link_tag = soup.find('a')
            if not link_tag: continue

            title = link_tag.get_text(strip=True)
            href = link_tag.get('href')
            full_url = BASE_URL + href if href.startswith('/') else href

            # 断点跳过
            if full_url in processed_urls:
                continue

            print(f"   Downloading: {title[:15]}...", end="")

            detail_data = extract_detail(page, full_url)

            if detail_data:
                if not detail_data['标题']: detail_data['标题'] = title
                save_row_immediately(detail_data, save_path)
                processed_urls.add(full_url)
                new_count += 1

            # 抓完详情页后，休息一下
            time.sleep(0.1)

        if len(matches) > 0 and new_count == 0:
            print("   (本页数据已全部存在，跳过)")
        elif new_count > 0:
            print(f"   (本页新增入库 {new_count} 条)")

    print(f"\n🎉 全部完成！")
    print(f"📁 文件位置: {save_path}")


if __name__ == "__main__":
    main()